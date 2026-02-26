import subprocess
import json
import os
from datetime import datetime
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from .models import Client, ClientNote, CustomFieldDefinition, ClientActivity, Provider, ClientFile, SystemSettings, DutySchedule, CustomHoliday, KktData, OfdCompany, OfdCompany
from .serializers import (
    ClientListSerializer, ClientDetailSerializer, ClientWriteSerializer,
    ClientNoteSerializer, CustomFieldDefinitionSerializer, ProviderSerializer, ClientFileSerializer,
    DutyScheduleSerializer, CustomHolidaySerializer, OfdCompanySerializer,
)
from apps.accounts.permissions import CanEditClient, CanManageCustomFields, IsAdmin


FIELD_LABELS = {
    'last_name': 'Фамилия',
    'first_name': 'Имя',
    'middle_name': 'Отчество',
    'inn': 'ИНН',
    'phone': 'Телефон',
    'iccid': 'ICCID',
    'email': 'Email',
    'pharmacy_code': 'Код аптеки',
    'company': 'Компания',
    'address': 'Адрес',
    'status': 'Статус',
    'provider_id': 'Провайдер',
    'personal_account': 'Лицевой счёт',
    'contract_number': '№ договора',
    'provider_settings': 'Настройки провайдера',
    'subnet': 'Подсеть аптеки',
    'external_ip': 'Внешний IP',
    'connection_type': 'Тип подключения',
    'tariff': 'Тариф',
    'modem_number': 'Номер модема/SIM',
    'modem_iccid': 'ICCID модема',
    'provider_equipment': 'Оборудование провайдера',
    'provider2_id': 'Провайдер 2',
    'personal_account2': 'Лицевой счёт 2',
    'contract_number2': '№ договора 2',
    'tariff2': 'Тариф 2',
    'connection_type2': 'Тип подключения 2',
    'modem_number2': 'Номер модема/SIM 2',
    'modem_iccid2': 'ICCID модема 2',
    'provider_equipment2': 'Оборудование провайдера 2',
}

STATUS_LABELS = {'active': 'Активен', 'inactive': 'Неактивен'}


def ping_ip(ip, timeout=5):
    try:
        result = subprocess.run(
            ['ping', '-c', '1', '-W', str(timeout), '-i', '0.2', ip],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=timeout + 1
        )
        return result.returncode == 0
    except Exception:
        return False


def build_change_log(old_client, new_data, provider_map):
    changes = []
    for field, label in FIELD_LABELS.items():
        old_val = getattr(old_client, field, '') or ''
        new_val = new_data.get(field.replace('_id', ''), '') or ''

        if field == 'provider_id':
            new_val = new_data.get('provider', '') or ''
            if str(old_val) == str(new_val):
                continue
            old_name = provider_map.get(old_val, f'#{old_val}') if old_val else '—'
            new_name = provider_map.get(int(new_val), f'#{new_val}') if new_val else '—'
            changes.append(f'{label}: «{old_name}» → «{new_name}»')
            continue

        if field == 'status':
            old_val = STATUS_LABELS.get(str(old_val), old_val)
            new_val = STATUS_LABELS.get(str(new_val), new_val)

        if field == 'provider_equipment':
            old_val = 'Присутствует' if str(old_val) == 'True' else 'Отсутствует'
            new_bool = str(new_data.get('provider_equipment', '')).lower()
            new_val = 'Присутствует' if new_bool in ('true', '1', 'yes') else 'Отсутствует'

        if str(old_val) != str(new_val):
            old_display = str(old_val)[:50] + '...' if len(str(old_val)) > 50 else str(old_val) or '—'
            new_display = str(new_val)[:50] + '...' if len(str(new_val)) > 50 else str(new_val) or '—'
            changes.append(f'{label}: «{old_display}» → «{new_display}»')
    return changes


class CustomFieldDefinitionViewSet(viewsets.ModelViewSet):
    queryset = CustomFieldDefinition.objects.filter(is_active=True)
    serializer_class = CustomFieldDefinitionSerializer
    permission_classes = [IsAuthenticated, CanManageCustomFields]

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return super().get_permissions()


class ClientViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, CanEditClient]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'provider']
    search_fields = ['company', 'address', 'phone', 'email', 'inn', 'pharmacy_code']
    ordering_fields = ['company', 'address', 'inn', 'phone', 'email', 'status', 'created_at', 'provider__name']
    ordering = ['-created_at']

    def get_queryset(self):
        # Удаляем черновики старше 2 часов (пользователь ушёл не сохранив)
        from django.utils import timezone
        from datetime import timedelta
        stale_drafts = Client.objects.filter(
            is_draft=True,
            created_at__lt=timezone.now() - timedelta(hours=2)
        )
        for draft in stale_drafts:
            for f in draft.files.all():
                f.file.delete()
                f.delete()
            draft.delete()

        qs = Client.objects.select_related('created_by', 'provider').filter(is_draft=False)
        providers_param = self.request.query_params.get('provider', '')
        if providers_param:
            provider_ids = [p.strip() for p in providers_param.split(',') if p.strip().isdigit()]
            if provider_ids:
                qs = qs.filter(provider__id__in=provider_ids)
        return qs

    def get_object_including_draft(self):
        from django.shortcuts import get_object_or_404
        return get_object_or_404(Client, pk=self.kwargs['pk'])

    def get_object(self):
        if self.action in ('update', 'partial_update', 'ping', 'files', 'delete_file', 'discard_draft'):
            from django.shortcuts import get_object_or_404
            obj = get_object_or_404(Client, pk=self.kwargs['pk'])
            self.check_object_permissions(self.request, obj)
            return obj
        return super().get_object()

    def get_serializer_class(self):
        if self.action == 'list':
            return ClientListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return ClientWriteSerializer
        return ClientDetailSerializer

    def perform_create(self, serializer):
        client = serializer.save(created_by=self.request.user)
        ClientActivity.objects.create(
            client=client, user=self.request.user,
            action='Карточка клиента создана'
        )

    def perform_update(self, serializer):
        old = self.get_object()
        if old.is_draft:
            # Конвертируем черновик в клиента только если явно передан finalize=true
            # или если запрос идёт не через saveDraftField (draft_save=true означает промежуточное сохранение)
            is_draft_save = self.request.data.get('_draft_save') == True or                             str(self.request.data.get('_draft_save', '')).lower() == 'true'
            if is_draft_save:
                # Промежуточное сохранение — оставляем черновиком
                serializer.save(is_draft=True)
                return
            client = serializer.save(is_draft=False)
            ClientActivity.objects.create(
                client=client, user=self.request.user,
                action='Карточка клиента создана'
            )
            return
        provider_map = {p.id: p.name for p in Provider.objects.all()}
        changes = build_change_log(old, self.request.data, provider_map)
        client = serializer.save()
        if changes:
            action = 'Изменено: ' + ' | '.join(changes)
            if len(action) > 490:
                action = action[:490] + '...'
        else:
            action = 'Карточка обновлена'
        ClientActivity.objects.create(client=client, user=self.request.user, action=action)

    def destroy(self, request, *args, **kwargs):
        if not request.user.has_perm_flag('can_delete_client'):
            return Response({'detail': 'Недостаточно прав.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get', 'post'])
    def notes(self, request, pk=None):
        client = self.get_object()
        if request.method == 'GET':
            return Response(ClientNoteSerializer(client.notes.all(), many=True).data)
        serializer = ClientNoteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        note = serializer.save(client=client, author=request.user)
        ClientActivity.objects.create(client=client, user=request.user, action='Добавлена заметка')
        return Response(ClientNoteSerializer(note).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get', 'post'], url_path='export_excel')
    def export_excel(self, request):
        import openpyxl
        import io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from .models import SystemSettings

        if request.method == 'POST':
            params = request.data
        else:
            params = request.query_params

        search = params.get('search', '').strip()
        status_filter = params.get('status', '').strip()
        send_via = params.get('send_via', 'file')
        to_email = params.get('to_email', '').strip()
        fields_raw = params.get('fields', '')
        if isinstance(fields_raw, list):
            selected_fields = set(fields_raw)
        else:
            selected_fields = set(f.strip() for f in fields_raw.split(',') if f.strip()) if fields_raw else None

        ALL_GROUPS = {'basic', 'network', 'provider1', 'provider2'}
        if not selected_fields:
            selected_fields = ALL_GROUPS

        qs = Client.objects.select_related('provider', 'provider2').filter(is_draft=False)
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(company__icontains=search) | Q(address__icontains=search) |
                Q(phone__icontains=search) | Q(email__icontains=search) |
                Q(inn__icontains=search) | Q(pharmacy_code__icontains=search)
            )
        if status_filter:
            qs = qs.filter(status=status_filter)
        clients = qs.order_by('-created_at')

        CONNECTION_LABELS = {
            'fiber': 'Оптоволокно', 'dsl': 'DSL', 'cable': 'Кабель',
            'wireless': 'Беспроводное', 'modem': 'Модем', 'mrnet': 'MR-Net',
        }
        STATUS_LABELS = {'active': 'Активен', 'inactive': 'Неактивен'}

        ALL_FIELDS = {
            'address':       ('Адрес',       35, lambda c: c.address or ''),
            'company':       ('Компания',    25, lambda c: c.company or ''),
            'inn':           ('ИНН',         14, lambda c: c.inn or ''),
            'phone':         ('Телефон',     16, lambda c: c.phone or ''),
            'email':         ('Email',       25, lambda c: c.email or ''),
            'pharmacy_code': ('Код аптеки',  12, lambda c: c.pharmacy_code or ''),
            'iccid':         ('ICCID',       22, lambda c: c.iccid or ''),
            'status':        ('Статус',      10, lambda c: STATUS_LABELS.get(c.status, c.status)),
            'subnet':        ('Подсеть',     16, lambda c: c.subnet or ''),
            'external_ip':   ('Внешний IP',  14, lambda c: c.external_ip or ''),
            'mikrotik_ip':   ('Микротик IP', 14, lambda c: c.mikrotik_ip or ''),
            'server_ip':     ('Сервер IP',   12, lambda c: c.server_ip or ''),
        }
        PROVIDER1_FIELDS = [
            ('Провайдер 1',               20, lambda c: c.provider.name if c.provider else ''),
            ('Тип подключения 1',         18, lambda c: CONNECTION_LABELS.get(c.connection_type, c.connection_type or '')),
            ('Тариф 1 (Мбит/с)',          14, lambda c: c.tariff or ''),
            ('Лицевой счёт 1',            16, lambda c: c.personal_account or ''),
            ('№ договора 1',              20, lambda c: c.contract_number or ''),
            ('Номер модема 1',            16, lambda c: c.modem_number or ''),
            ('ICCID модема 1',            22, lambda c: c.modem_iccid or ''),
            ('Оборудование провайдера 1', 12, lambda c: 'Присутствует' if c.provider_equipment else 'Отсутствует'),
        ]
        PROVIDER2_FIELDS = [
            ('Провайдер 2',               20, lambda c: c.provider2.name if c.provider2 else ''),
            ('Тип подключения 2',         18, lambda c: CONNECTION_LABELS.get(c.connection_type2, c.connection_type2 or '')),
            ('Тариф 2 (Мбит/с)',          14, lambda c: c.tariff2 or ''),
            ('Лицевой счёт 2',            16, lambda c: c.personal_account2 or ''),
            ('№ договора 2',              20, lambda c: c.contract_number2 or ''),
            ('Номер модема 2',            16, lambda c: c.modem_number2 or ''),
            ('ICCID модема 2',            22, lambda c: c.modem_iccid2 or ''),
            ('Оборудование провайдера 2', 12, lambda c: 'Присутствует' if c.provider_equipment2 else 'Отсутствует'),
        ]

        FIELD_ORDER = [
            'address', 'company', 'inn', 'phone', 'email',
            'pharmacy_code', 'iccid', 'status',
            'subnet', 'external_ip', 'mikrotik_ip', 'server_ip',
        ]

        headers = []
        for key in FIELD_ORDER:
            if key in selected_fields:
                headers.append(ALL_FIELDS[key])
        if 'provider1' in selected_fields:
            headers += PROVIDER1_FIELDS
        if 'provider2' in selected_fields:
            headers += PROVIDER2_FIELDS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Клиенты'

        header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', start_color='1677FF')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align   = Alignment(vertical='center', wrap_text=True)
        thin   = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, (title, width, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 35

        for row, c in enumerate(clients, 2):
            fill = PatternFill('solid', start_color='F8FBFF') if row % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
            for col, (_, _, getter) in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=getter(c))
                cell.font      = Font(name='Arial', size=10)
                cell.alignment = cell_align
                cell.border    = border
                cell.fill      = fill
            ws.row_dimensions[row].height = 20

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}1'

        filename = f'clients_{__import__("datetime").date.today()}.xlsx'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        excel_bytes = buf.read()

        if send_via == 'email':
            if not to_email:
                return Response({'error': 'Не указан email для отправки'}, status=400)
            s = SystemSettings.get()
            if not s.smtp_host or not s.smtp_user or not s.smtp_password_encrypted or not s.smtp_from_email:
                return Response({'error': 'SMTP настройки не заполнены'}, status=400)
            try:
                import smtplib, ssl
                from email.mime.multipart import MIMEMultipart
                from email.mime.base import MIMEBase
                from email.mime.text import MIMEText
                from email import encoders

                from_addr = f'{s.smtp_from_name} <{s.smtp_from_email}>' if s.smtp_from_name else s.smtp_from_email
                msg = MIMEMultipart()
                msg['Subject'] = f'Выгрузка клиентов — {__import__("datetime").date.today()}'
                msg['From']    = from_addr
                msg['To']      = to_email
                msg.attach(MIMEText(
                    f'<html><body style="font-family:Arial,sans-serif;padding:20px;">'
                    f'<h2 style="color:#1677ff;">Выгрузка клиентов</h2>'
                    f'<p>Во вложении файл Excel с выгрузкой клиентов на {__import__("datetime").date.today()}.</p>'
                    f'<p style="color:#999;font-size:12px;">Support Portal</p>'
                    f'</body></html>', 'html', 'utf-8'
                ))
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(excel_bytes)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)

                if s.smtp_use_ssl:
                    ctx = ssl.create_default_context()
                    with smtplib.SMTP_SSL(s.smtp_host, s.smtp_port, context=ctx, timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password)
                        srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                elif s.smtp_use_tls:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.ehlo(); srv.starttls(); srv.ehlo()
                        srv.login(s.smtp_user, s.smtp_password)
                        srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                else:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password)
                        srv.sendmail(s.smtp_from_email, to_email, msg.as_string())

                return Response({'success': True, 'message': f'Файл отправлен на {to_email}'})
            except smtplib.SMTPAuthenticationError:
                return Response({'error': 'Ошибка аутентификации SMTP'}, status=400)
            except Exception as e:
                return Response({'error': f'Ошибка отправки: {str(e)}'}, status=400)

        response = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='create_draft')
    def create_draft(self, request):
        from django.utils import timezone
        from datetime import timedelta
        old_drafts = Client.objects.filter(
            is_draft=True,
            created_by=request.user,
            created_at__lt=timezone.now() - timedelta(hours=1)
        )
        for draft in old_drafts:
            for f in draft.files.all():
                f.file.delete()
                f.delete()
            draft.delete()
        client = Client.objects.create(
            is_draft=True,
            created_by=request.user,
            last_name='', first_name='',
        )
        return Response({'id': client.id}, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete', 'post'], url_path='discard_draft')
    def discard_draft(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk, is_draft=True, created_by=request.user)
            for f in client.files.all():
                f.file.delete()
                f.delete()
            client.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Client.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get', 'post'], url_path='files')
    def files(self, request, pk=None):
        client = self.get_object_including_draft()
        if request.method == 'GET':
            files = client.files.all()
            return Response(ClientFileSerializer(files, many=True, context={'request': request}).data)
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Файл не передан.'}, status=status.HTTP_400_BAD_REQUEST)
        if file.size > 5 * 1024 * 1024:
            return Response({'detail': 'Файл слишком большой. Максимум 5 МБ.'}, status=status.HTTP_400_BAD_REQUEST)
        client_file = ClientFile.objects.create(
            client=client, file=file, name=file.name,
            size=file.size, uploaded_by=request.user,
        )
        ClientActivity.objects.create(
            client=client, user=request.user,
            action=f'Загружен файл: {file.name}'
        )
        return Response(
            ClientFileSerializer(client_file, context={'request': request}).data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['delete'], url_path='files/(?P<file_id>[^/.]+)')
    def delete_file(self, request, pk=None, file_id=None):
        client = self.get_object_including_draft()
        try:
            client_file = ClientFile.objects.get(id=file_id, client=client)
            name = client_file.name
            client_file.file.delete()
            client_file.delete()
            ClientActivity.objects.create(
                client=client, user=request.user,
                action=f'Удалён файл: {name}'
            )
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ClientFile.DoesNotExist:
            return Response({'detail': 'Файл не найден.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='transfer_modem')
    def transfer_modem(self, request, pk=None):
        from_client = self.get_object()
        to_client_id = request.data.get('to_client_id')
        from_slot = str(request.data.get('from_slot', '1'))
        to_slot = str(request.data.get('to_slot', '1'))

        if not to_client_id:
            return Response({'error': 'Не указан клиент для передачи'}, status=400)

        try:
            to_client = Client.objects.get(pk=to_client_id, is_draft=False)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=400)

        if from_client.pk == to_client.pk:
            return Response({'error': 'Нельзя передать самому себе'}, status=400)

        from_sfx = '' if from_slot == '1' else '2'
        to_sfx   = '' if to_slot   == '1' else '2'

        PROVIDER_FIELDS = [
            'provider', 'personal_account', 'contract_number', 'tariff',
            'connection_type', 'modem_number', 'modem_iccid',
            'provider_settings', 'provider_equipment',
        ]

        from_data = {}
        for f in PROVIDER_FIELDS:
            from_data[f] = getattr(from_client, f + from_sfx, None)

        for f in PROVIDER_FIELDS:
            setattr(to_client, f + to_sfx, from_data[f])
        to_client.save()

        prov_name = (from_data['provider'].name if from_data['provider'] else '—')
        from_name = from_client.company or from_client.address or f'Клиент #{from_client.pk}'
        to_name   = to_client.company   or to_client.address   or f'Клиент #{to_client.pk}'
        conn_labels = {
            'fiber': 'Оптоволокно', 'dsl': 'DSL', 'cable': 'Кабель',
            'wireless': 'Беспроводное', 'modem': 'Модем', 'mrnet': 'MR-Net',
        }
        conn_label = conn_labels.get(from_data['connection_type'] or '', from_data['connection_type'] or '—')
        equip_label = 'Присутствует' if from_data['provider_equipment'] else 'Отсутствует'

        log_fields = (
            f'Провайдер: {prov_name}\n'
            f'Тип: {conn_label}\n'
            f'Тариф: {from_data["tariff"] or "—"} Мбит/с\n'
            f'Лицевой счёт: {from_data["personal_account"] or "—"}\n'
            f'№ договора: {from_data["contract_number"] or "—"}\n'
            f'Номер модема: {from_data["modem_number"] or "—"}\n'
            f'ICCID модема: {from_data["modem_iccid"] or "—"}\n'
            f'Оборудование: {equip_label}'
        )

        ClientActivity.objects.create(
            client=to_client, user=request.user,
            action=(
                f'Получен провайдер {to_slot} от клиента «{from_name}»:\n{log_fields}'
            )
        )

        EMPTY_VALUES = {
            'provider': None, 'personal_account': '', 'contract_number': '', 'tariff': '',
            'connection_type': '', 'modem_number': '', 'modem_iccid': '',
            'provider_settings': '', 'provider_equipment': False,
        }
        for f, empty in EMPTY_VALUES.items():
            setattr(from_client, f + from_sfx, empty)
        from_client.save()

        ClientActivity.objects.create(
            client=from_client, user=request.user,
            action=(
                f'Провайдер {from_slot} передан клиенту «{to_name}» в слот {to_slot}:\n{log_fields}'
            )
        )

        return Response({
            'success': True,
            'to_client': {'id': to_client.pk, 'name': to_client.company or to_client.address or f'#{to_client.pk}'},
        })

    @action(detail=True, methods=['get'], url_path='ping')
    def ping(self, request, pk=None):
        client = self.get_object_including_draft()
        external_ip = client.external_ip or ''
        mikrotik_ip = client.mikrotik_ip or ''
        server_ip = client.server_ip or ''
        results = {
            'external_ip': {'ip': external_ip, 'alive': ping_ip(external_ip) if external_ip else None},
            'mikrotik_ip': {'ip': mikrotik_ip, 'alive': ping_ip(mikrotik_ip) if mikrotik_ip else None},
            'server_ip': {'ip': server_ip, 'alive': ping_ip(server_ip) if server_ip else None},
        }
        return Response(results)


class FetchExternalIPView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import paramiko
        mikrotik_ip = request.data.get('mikrotik_ip', '').strip()
        old_external_ip = request.data.get('old_external_ip', '').strip()

        if not mikrotik_ip:
            return Response({'error': 'Микротик IP не передан'}, status=400)

        settings_obj = SystemSettings.get()
        if not settings_obj.ssh_user:
            return Response({'error': 'SSH пользователь не задан в настройках системы'}, status=400)
        if not settings_obj.ssh_password_encrypted:
            return Response({'error': 'SSH пароль не задан в настройках системы'}, status=400)

        if not ping_ip(mikrotik_ip):
            return Response({'error': f'Микротик {mikrotik_ip} недоступен'}, status=400)

        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                mikrotik_ip,
                username=settings_obj.ssh_user,
                password=settings_obj.ssh_password,
                timeout=10, port=22
            )
            cmd = ':put ([/tool fetch url="https://api.ipify.org" http-method=get output=user as-value]->"data")'
            stdin, stdout, stderr = ssh.exec_command(cmd, timeout=15)
            new_ip = stdout.read().decode().strip()
            ssh.close()

            if not new_ip:
                return Response({'error': 'Не удалось получить внешний IP с Микротика'}, status=400)

            return Response({
                'new_ip': new_ip,
                'old_ip': old_external_ip,
                'changed': new_ip != old_external_ip,
            })
        except paramiko.AuthenticationException:
            return Response({'error': f'Ошибка аутентификации SSH на {mikrotik_ip}'}, status=400)
        except Exception as e:
            return Response({'error': f'Ошибка подключения к {mikrotik_ip}: {str(e)}'}, status=400)


class OfdCompanyViewSet(viewsets.ModelViewSet):
    queryset = OfdCompany.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return OfdCompanyWriteSerializer
        return OfdCompanySerializer


class ProviderViewSet(viewsets.ModelViewSet):
    queryset = Provider.objects.all()
    serializer_class = ProviderSerializer
    permission_classes = [IsAuthenticated, CanEditClient]


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Count
        from django.utils import timezone
        from datetime import timedelta

        clients = Client.objects.filter(is_draft=False)
        now = timezone.now()
        month_ago = now - timedelta(days=30)
        week_ago = now - timedelta(days=7)

        total = clients.count()
        active = clients.filter(status='active').count()
        inactive = clients.filter(status='inactive').count()
        new_month = clients.filter(created_at__gte=month_ago).count()
        new_week = clients.filter(created_at__gte=week_ago).count()

        by_provider = list(
            clients.filter(provider__isnull=False)
            .values('provider__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )
        by_provider_data = [{'name': r['provider__name'], 'value': r['count']} for r in by_provider]
        no_provider = clients.filter(provider__isnull=True).count()
        if no_provider > 0:
            by_provider_data.append({'name': 'Без провайдера', 'value': no_provider})

        conn_labels = {
            'fiber': 'Оптоволокно', 'dsl': 'DSL', 'cable': 'Кабель',
            'wireless': 'Беспроводное', 'modem': 'Модем', 'mrnet': 'MR-Net',
        }
        by_connection = list(
            clients.filter(connection_type__isnull=False)
            .exclude(connection_type='')
            .values('connection_type')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        by_connection_data = [
            {'name': conn_labels.get(r['connection_type'], r['connection_type']), 'value': r['count']}
            for r in by_connection
        ]

        monthly = []
        for i in range(5, -1, -1):
            d = now - timedelta(days=30 * i)
            month_start = d.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if i > 0:
                next_month = (d + timedelta(days=32)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                next_month = now
            cnt = clients.filter(created_at__gte=month_start, created_at__lt=next_month).count()
            month_names = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн',
                           'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
            monthly.append({'month': month_names[month_start.month - 1], 'count': cnt})

        recent_activities = []
        for a in ClientActivity.objects.select_related('user', 'client').order_by('-created_at')[:10]:
            recent_activities.append({
                'id': a.id,
                'action': a.action,
                'user_name': a.user.full_name or a.user.username if a.user else '—',
                'client_id': a.client_id,
                'client_name': a.client.address or a.client.company or f'Клиент #{a.client_id}',
                'created_at': a.created_at.isoformat(),
            })

        return Response({
            'total': total,
            'active': active,
            'inactive': inactive,
            'new_month': new_month,
            'new_week': new_week,
            'by_provider': by_provider_data,
            'by_connection': by_connection_data,
            'monthly': monthly,
            'recent_activities': recent_activities,
        })


class DutyScheduleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = DutyScheduleSerializer

    def get_queryset(self):
        qs = DutySchedule.objects.select_related('user')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)
        return qs

    @action(detail=False, methods=['post'])
    def clear_month(self, request):
        if not (request.user.is_superuser or (request.user.role and request.user.role.name == 'admin')):
            return Response({'error': 'Недостаточно прав'}, status=403)
        year = request.data.get('year')
        month = request.data.get('month')
        if not year or not month:
            return Response({'error': 'year и month обязательны'}, status=400)
        deleted, _ = DutySchedule.objects.filter(date__year=year, date__month=month).delete()
        return Response({'deleted': deleted})

    @action(detail=False, methods=['get'])
    def holidays(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        qs = CustomHoliday.objects.all()
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)
        return Response(CustomHolidaySerializer(qs, many=True).data)

    @action(detail=False, methods=['post'])
    def toggle_holiday(self, request):
        date = request.data.get('date')
        is_holiday = request.data.get('is_holiday')
        note = request.data.get('note', '')
        if not date:
            return Response({'error': 'date обязателен'}, status=400)
        if is_holiday is None:
            CustomHoliday.objects.filter(date=date).delete()
            return Response({'status': 'deleted'})
        obj, _ = CustomHoliday.objects.update_or_create(
            date=date,
            defaults={'is_holiday': is_holiday, 'note': note, 'created_by': request.user}
        )
        return Response(CustomHolidaySerializer(obj).data)

    @action(detail=False, methods=['post'])
    def set_duty(self, request):
        user_id = request.data.get('user_id')
        date = request.data.get('date')
        duty_type = request.data.get('duty_type')

        if not user_id or not date:
            return Response({'error': 'user_id и date обязательны'}, status=400)

        if duty_type is None or duty_type == '':
            DutySchedule.objects.filter(user_id=user_id, date=date).delete()
            return Response({'status': 'deleted'})

        obj, created = DutySchedule.objects.update_or_create(
            user_id=user_id, date=date,
            defaults={'duty_type': duty_type}
        )
        return Response(DutyScheduleSerializer(obj).data)

    @action(detail=False, methods=['get'])
    def report(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        qs = DutySchedule.objects.select_related('user')
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)

        from apps.accounts.models import User

        users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
        duty_types = ['phone', 'day', 'phone_day', 'vacation', 'busy']
        labels = {
            'phone': 'Телефон',
            'day': 'Работа днём',
            'phone_day': 'Телефон + день',
            'vacation': 'Отпуск',
            'busy': 'Занят',
        }

        report_data = []
        for u in users:
            row = {
                'user_id': u.id,
                'user_name': u.full_name or u.email,
                'totals': {},
            }
            for dt in duty_types:
                row['totals'][dt] = qs.filter(user=u, duty_type=dt).count()
            row['total'] = sum(row['totals'].values())
            report_data.append(row)

        return Response({'report': report_data, 'labels': labels})


# ===================== Компании ОФД =====================

class OfdCompanyViewSet(viewsets.ModelViewSet):
    queryset = OfdCompany.objects.all()
    serializer_class = OfdCompanySerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsAuthenticated()]
        return [IsAuthenticated()]


# ===================== ОФД / ККТ =====================

def parse_datetime(value):
    """Парсит дату из строки ОФД (ISO 8601 без timezone)"""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def ofd_request(inn, token, search=''):
    """
    Один вызов скрипта — получает все ККТ по ИНН, фильтрует по адресу,
    возвращает детали по каждому РНМ. Если search — 16 цифр, делает прямой запрос по РНМ.
    """
    script = '/usr/local/bin/ofd_fetch.sh'
    try:
        args = [script, inn, token, search]
        result = subprocess.run(args, capture_output=True, text=True, timeout=300)
        if not result.stdout.strip():
            stderr = result.stderr.strip()
            raise Exception(f'Пустой ответ от скрипта. stderr: {stderr}')
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        raise Exception(f'Некорректный JSON: {result.stdout[:300]}')
    except subprocess.TimeoutExpired:
        raise Exception('Превышено время ожидания (300 сек)')
    except FileNotFoundError:
        raise Exception(f'Скрипт {script} не найден в контейнере')


class OfdKktView(APIView):
    """
    GET   /api/clients/{id}/ofd_kkt/  — получить сохранённые ККТ из БД
    POST  /api/clients/{id}/ofd_kkt/  — запросить данные с ОФД (полный поиск по ИНН+адресу)
    PATCH /api/clients/{id}/ofd_kkt/  — быстрое обновление по сохранённым РНМ
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        kkts = client.kkt_data.all()
        data = []
        for k in kkts:
            data.append({
                'id': k.id,
                'kkt_reg_id': k.kkt_reg_id,
                'serial_number': k.serial_number,
                'fn_number': k.fn_number,
                'kkt_model': k.kkt_model,
                'create_date': k.create_date.isoformat() if k.create_date else None,
                'check_date': k.check_date.isoformat() if k.check_date else None,
                'activation_date': k.activation_date.isoformat() if k.activation_date else None,
                'first_document_date': k.first_document_date.isoformat() if k.first_document_date else None,
                'contract_start_date': k.contract_start_date.isoformat() if k.contract_start_date else None,
                'contract_end_date': k.contract_end_date.isoformat() if k.contract_end_date else None,
                'fn_end_date': k.fn_end_date.isoformat() if k.fn_end_date else None,
                'last_doc_on_kkt': k.last_doc_on_kkt.isoformat() if k.last_doc_on_kkt else None,
                'last_doc_on_ofd': k.last_doc_on_ofd.isoformat() if k.last_doc_on_ofd else None,
                'fiscal_address': k.fiscal_address,
                'fetched_at': k.fetched_at.isoformat() if k.fetched_at else None,
            })
        return Response(data)

    def post(self, request, pk=None):
        """Полный поиск: по ИНН компании + фильтрация по адресу аптеки"""
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        if not client.ofd_company:
            return Response({'error': 'У клиента не выбрана компания. Укажите компанию в карточке клиента.'}, status=400)

        address = (client.address or '').strip()
        if not address:
            return Response({'error': 'У клиента не заполнен адрес'}, status=400)

        inn = client.ofd_company.inn.strip()
        token = client.ofd_company.ofd_token.strip()

        if not inn:
            return Response({'error': f'У компании «{client.ofd_company.name}» не заполнен ИНН'}, status=400)
        if not token:
            return Response({'error': f'У компании «{client.ofd_company.name}» не задан токен ОФД. Откройте раздел «Компании» и добавьте токен.'}, status=400)

        # Один вызов скрипта — он сам фильтрует по адресу и получает детали
        try:
            import time
            t_start = time.time()
            result_data = ofd_request(inn, token, address)
            elapsed = round(time.time() - t_start, 1)
        except Exception as e:
            return Response({'error': f'Ошибка запроса к ОФД: {str(e)}'}, status=502)

        if result_data.get('Status') != 'Success':
            return Response({'error': f'ОФД вернул ошибку: {result_data}'}, status=502)

        all_items = result_data.get('Data', [])
        if not all_items:
            return Response({'error': 'ОФД не вернул ни одной ККТ для данного ИНН и адреса'}, status=404)

        fetched = []
        errors = []
        for item in all_items:
            kkt_obj, _ = KktData.objects.get_or_create(
                client=client,
                kkt_reg_id=item.get('KktRegId', ''),
            )
            kkt_obj.serial_number = item.get('SerialNumber', '')
            kkt_obj.fn_number = item.get('FnNumber', '')
            kkt_obj.kkt_model = item.get('KktModel', '')
            kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
            kkt_obj.check_date = parse_datetime(item.get('CheckDate'))
            kkt_obj.activation_date = parse_datetime(item.get('ActivationDate'))
            kkt_obj.first_document_date = parse_datetime(item.get('FirstDocumentDate'))
            kkt_obj.contract_start_date = parse_datetime(item.get('ContractStartDate'))
            kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
            kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
            kkt_obj.last_doc_on_kkt = parse_datetime(item.get('LastDocOnKktDateTime'))
            kkt_obj.last_doc_on_ofd = parse_datetime(item.get('LastDocOnOfdDateTimeUtc'))
            kkt_obj.fiscal_address = item.get('FiscalAddress', '')
            kkt_obj.raw_data = item
            kkt_obj.save()
            fetched.append(kkt_obj.kkt_reg_id)

        return Response({
            'success': True,
            'fetched': fetched,
            'errors': errors,
            'elapsed': elapsed,
            'message': f'Получено ККТ: {len(fetched)}, время: {elapsed} сек' + (f', ошибок: {len(errors)}' if errors else ''),
        })

    def delete(self, request, pk=None, kkt_id=None):
        """DELETE /api/clients/{id}/ofd_kkt/{kkt_id}/ — удалить одну ККТ"""
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        if not kkt_id:
            return Response({'error': 'Не указан ID ККТ'}, status=400)

        try:
            kkt = client.kkt_data.get(pk=kkt_id)
            kkt.delete()
            return Response({'success': True, 'message': 'ККТ удалена'})
        except KktData.DoesNotExist:
            return Response({'error': 'ККТ не найдена'}, status=404)

    def patch(self, request, pk=None):
        """
        Обновление по РНМ.
        Если в теле запроса передан rnm_override — используем его (режим создания клиента).
        Иначе берём РНМ из уже сохранённых в БД записей KktData.
        """
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        if not client.ofd_company:
            return Response({'error': 'У клиента не выбрана компания'}, status=400)

        inn = client.ofd_company.inn.strip()
        token = client.ofd_company.ofd_token.strip()

        if not inn:
            return Response({'error': f'У компании «{client.ofd_company.name}» не заполнен ИНН'}, status=400)
        if not token:
            return Response({'error': f'У компании «{client.ofd_company.name}» не задан токен ОФД'}, status=400)

        # Если передан конкретный РНМ — работаем только с ним (для режима создания)
        rnm_override = (request.data.get('rnm_override') or '').strip()
        if rnm_override:
            import re
            if not re.match(r'^\d{16}$', rnm_override):
                return Response({'error': f'Некорректный РНМ: «{rnm_override}». Должно быть 16 цифр.'}, status=400)
            import time
            t_start = time.time()
            script = '/usr/local/bin/ofd_fetch.sh'
            try:
                result = subprocess.run(
                    [script, inn, token, rnm_override],
                    capture_output=True, text=True, timeout=60
                )
                if not result.stdout.strip():
                    return Response({'error': f'Пустой ответ от скрипта для РНМ {rnm_override}. stderr: {result.stderr[:200]}'}, status=502)
                detail_data = json.loads(result.stdout)
            except Exception as e:
                return Response({'error': f'Ошибка запроса для РНМ {rnm_override}: {str(e)}'}, status=502)

            if detail_data.get('Status') != 'Success':
                return Response({'error': f'ОФД вернул ошибку для РНМ {rnm_override}'}, status=502)

            items = detail_data.get('Data', [])
            if not items:
                return Response({'error': f'ОФД не вернул данных для РНМ {rnm_override}'}, status=404)

            # Проверяем что адрес ККТ совпадает с адресом клиента
            client_address = (client.address or '').strip()
            if client_address:
                import re as _re
                stop_words = {'ул', 'пр', 'пом', 'д', 'кв', 'г', 'р-н', 'ул.', 'пр.', 'д.', 'пом.',
                              'г.', 'пр-кт', 'пр-кт.', 'проспект', 'улица', 'переулок', 'помещ',
                              'помещение', 'область', 'край', 'республика', 'город'}
                words = [w.strip('.,()-').lower() for w in _re.split(r'[\s,]+', client_address)]
                words = [w for w in words if w and len(w) > 1 and w not in stop_words]
                house_numbers = _re.findall(r'\b(\d+)\b', client_address)
                house_num = house_numbers[-1] if house_numbers else None

                fiscal_addr = (items[0].get('FiscalAddress') or '').lower()
                addr_numbers = _re.findall(r'\b(\d+)\b', items[0].get('FiscalAddress', ''))

                word_match = all(w in fiscal_addr for w in words)
                house_match = (not house_num) or (house_num in addr_numbers)

                if not (word_match and house_match):
                    return Response({
                        'success': False,
                        'fetched': [],
                        'errors': [
                            f'РНМ {rnm_override}: адрес ККТ в ОФД («{items[0].get("FiscalAddress", "")}») '
                            f'не совпадает с адресом клиента («{client_address}»)'
                        ],
                        'elapsed': round(time.time() - t_start, 1),
                        'message': f'РНМ {rnm_override} не добавлен: адрес не совпадает',
                    }, status=200)  # 200 чтобы фронт показал предупреждение, не ошибку

            for item in items:
                kkt_obj, _ = KktData.objects.get_or_create(
                    client=client,
                    kkt_reg_id=item.get('KktRegId', rnm_override),
                )
                kkt_obj.serial_number = item.get('SerialNumber', '')
                kkt_obj.fn_number = item.get('FnNumber', '')
                kkt_obj.kkt_model = item.get('KktModel', '')
                kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
                kkt_obj.check_date = parse_datetime(item.get('CheckDate'))
                kkt_obj.activation_date = parse_datetime(item.get('ActivationDate'))
                kkt_obj.first_document_date = parse_datetime(item.get('FirstDocumentDate'))
                kkt_obj.contract_start_date = parse_datetime(item.get('ContractStartDate'))
                kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
                kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
                kkt_obj.last_doc_on_kkt = parse_datetime(item.get('LastDocOnKktDateTime'))
                kkt_obj.last_doc_on_ofd = parse_datetime(item.get('LastDocOnOfdDateTimeUtc'))
                kkt_obj.fiscal_address = item.get('FiscalAddress', '')
                kkt_obj.raw_data = item
                kkt_obj.save()

            elapsed = round(time.time() - t_start, 1)
            return Response({
                'success': True,
                'fetched': [rnm_override],
                'errors': [],
                'elapsed': elapsed,
                'message': f'Получены данные по РНМ {rnm_override}, время: {elapsed} сек',
            })

        # Стандартный режим — обновляем по РНМ из БД
        kkts = client.kkt_data.all()
        if not kkts:
            return Response({'error': 'Нет сохранённых ККТ. Сначала нажмите «Получить данные с ОФД»'}, status=404)

        import time
        t_start = time.time()
        script = '/usr/local/bin/ofd_fetch.sh'
        fetched = []
        errors = []

        for kkt_obj in kkts:
            rnm = kkt_obj.kkt_reg_id
            if not rnm:
                continue
            try:
                result = subprocess.run(
                    [script, inn, token, rnm],
                    capture_output=True, text=True, timeout=60
                )
                if not result.stdout.strip():
                    errors.append(f'РНМ {rnm}: пустой ответ')
                    continue
                detail_data = json.loads(result.stdout)
            except Exception as e:
                errors.append(f'РНМ {rnm}: {str(e)}')
                continue

            if detail_data.get('Status') != 'Success':
                errors.append(f'РНМ {rnm}: ОФД вернул ошибку')
                continue

            for item in detail_data.get('Data', []):
                kkt_obj.serial_number = item.get('SerialNumber', '')
                kkt_obj.fn_number = item.get('FnNumber', '')
                kkt_obj.kkt_model = item.get('KktModel', '')
                kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
                kkt_obj.check_date = parse_datetime(item.get('CheckDate'))
                kkt_obj.activation_date = parse_datetime(item.get('ActivationDate'))
                kkt_obj.first_document_date = parse_datetime(item.get('FirstDocumentDate'))
                kkt_obj.contract_start_date = parse_datetime(item.get('ContractStartDate'))
                kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
                kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
                kkt_obj.last_doc_on_kkt = parse_datetime(item.get('LastDocOnKktDateTime'))
                kkt_obj.last_doc_on_ofd = parse_datetime(item.get('LastDocOnOfdDateTimeUtc'))
                kkt_obj.fiscal_address = item.get('FiscalAddress', '')
                kkt_obj.raw_data = item
                kkt_obj.save()
                fetched.append(rnm)

        elapsed = round(time.time() - t_start, 1)
        return Response({
            'success': True,
            'fetched': fetched,
            'errors': errors,
            'elapsed': elapsed,
            'message': f'Обновлено ККТ: {len(fetched)}, время: {elapsed} сек' + (f', ошибок: {len(errors)}' if errors else ''),
        })
