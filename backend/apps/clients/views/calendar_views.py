import subprocess
import json
import os
from datetime import datetime
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from ..models import Client, ClientNote, CustomFieldDefinition, ClientActivity, Provider, ClientFile, SystemSettings, DutySchedule, CustomHoliday, KktData, OfdCompany
from ..serializers import (
    ClientListSerializer, ClientDetailSerializer, ClientWriteSerializer,
    ClientNoteSerializer, CustomFieldDefinitionSerializer, ProviderSerializer, ClientFileSerializer,
    DutyScheduleSerializer, CustomHolidaySerializer, OfdCompanySerializer, OfdCompanyWriteSerializer,
)
from apps.accounts.permissions import CanEditClient, CanManageCustomFields, IsAdmin
from .utils import ping_ip, build_change_log, FIELD_LABELS, STATUS_LABELS


class DutyScheduleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = DutyScheduleSerializer
    pagination_class = None  # Отключаем пагинацию — нужны все записи месяца

    def get_queryset(self):
        qs = DutySchedule.objects.select_related('user')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)
        return qs

    @action(detail=False, methods=['post'])
    def clear_month(self, request):
        if not (request.user.is_superuser or (request.user.role and request.user.role.name == 'admin')):
            return Response({'error': 'Недостаточно прав'}, status=403)
        year = request.data.get('year')
        month = request.data.get('month')
        if not year or not month:
            return Response({'error': 'year и month обязательны'}, status=400)
        deleted, _ = DutySchedule.objects.filter(date__year=year, date__month=month).delete()
        return Response({'deleted': deleted})

    @action(detail=False, methods=['get'])
    def holidays(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        qs = CustomHoliday.objects.all()
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)
        return Response(CustomHolidaySerializer(qs, many=True).data)

    @action(detail=False, methods=['post'])
    def toggle_holiday(self, request):
        date = request.data.get('date')
        is_holiday = request.data.get('is_holiday')
        note = request.data.get('note', '')
        if not date:
            return Response({'error': 'date обязателен'}, status=400)
        if is_holiday is None:
            CustomHoliday.objects.filter(date=date).delete()
            return Response({'status': 'deleted'})
        obj, _ = CustomHoliday.objects.update_or_create(
            date=date,
            defaults={'is_holiday': is_holiday, 'note': note, 'created_by': request.user}
        )
        return Response(CustomHolidaySerializer(obj).data)

    @action(detail=False, methods=['post'])
    def set_duty(self, request):
        user_id = request.data.get('user_id')
        date = request.data.get('date')
        duty_type = request.data.get('duty_type')

        if not user_id or not date:
            return Response({'error': 'user_id и date обязательны'}, status=400)

        if duty_type is None or duty_type == '':
            DutySchedule.objects.filter(user_id=user_id, date=date).delete()
            return Response({'status': 'deleted'})

        obj, created = DutySchedule.objects.update_or_create(
            user_id=user_id, date=date,
            defaults={'duty_type': duty_type}
        )
        return Response(DutyScheduleSerializer(obj).data)

    @action(detail=False, methods=['post'], url_path='bulk_set_duty')
    def bulk_set_duty(self, request):
        """Массовое назначение дежурства.
        Режим 1 (один сотрудник): user_id + dates[]
        Режим 2 (любые ячейки):   cells: [{user_id, date}, ...]
        """
        duty_type = request.data.get('duty_type')
        cells_raw = request.data.get('cells')

        if cells_raw is not None:
            # Режим 2 — произвольные ячейки
            if not isinstance(cells_raw, list) or len(cells_raw) == 0:
                return Response({'error': 'cells должен быть непустым списком'}, status=400)
            pairs = [(c.get('user_id'), c.get('date')) for c in cells_raw if c.get('user_id') and c.get('date')]
            if not pairs:
                return Response({'error': 'Нет валидных ячеек'}, status=400)

            if duty_type is None or duty_type == '':
                for user_id, date in pairs:
                    DutySchedule.objects.filter(user_id=user_id, date=date).delete()
            else:
                for user_id, date in pairs:
                    DutySchedule.objects.update_or_create(
                        user_id=user_id, date=date,
                        defaults={'duty_type': duty_type}
                    )
            return Response({'status': 'ok', 'count': len(pairs)})

        # Режим 1 — один сотрудник, несколько дат
        user_id = request.data.get('user_id')
        dates   = request.data.get('dates', [])
        if not user_id or not dates:
            return Response({'error': 'user_id и dates обязательны'}, status=400)

        if duty_type is None or duty_type == '':
            DutySchedule.objects.filter(user_id=user_id, date__in=dates).delete()
        else:
            for date in dates:
                DutySchedule.objects.update_or_create(
                    user_id=user_id, date=date,
                    defaults={'duty_type': duty_type}
                )
        return Response({'status': 'ok', 'count': len(dates)})

    @action(detail=False, methods=['get'])
    def report(self, request):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        qs = DutySchedule.objects.select_related('user')
        if year and month:
            qs = qs.filter(date__year=year, date__month=month)

        from apps.accounts.models import User

        users = User.objects.filter(is_active=True).order_by('last_name', 'first_name')
        duty_types = ['phone', 'day', 'phone_day', 'vacation', 'busy']
        labels = {
            'phone': 'Телефон',
            'day': 'Работа днём',
            'phone_day': 'Телефон + день',
            'vacation': 'Отпуск',
            'busy': 'Занят',
        }

        report_data = []
        for u in users:
            row = {
                'user_id': u.id,
                'user_name': u.full_name or u.email,
                'totals': {},
            }
            for dt in duty_types:
                row['totals'][dt] = qs.filter(user=u, duty_type=dt).count()
            row['total'] = sum(row['totals'].values())
            report_data.append(row)

        return Response({'report': report_data, 'labels': labels})

    @action(detail=False, methods=['get'], url_path='vacation_report')
    def vacation_report(self, request):
        """График отпусков за год: сотрудники × месяцы, периоды и пересечения."""
        import datetime as dt
        from apps.accounts.models import User

        year = int(request.query_params.get('year', dt.date.today().year))
        users = User.objects.filter(is_active=True).exclude(
            role__name='communications'
        ).order_by('last_name', 'first_name')

        # Все отпуска за год
        vacations = DutySchedule.objects.filter(
            duty_type='vacation',
            date__year=year,
        ).select_related('user').order_by('date')

        # Группируем по пользователю → список дат
        user_dates = {}
        for v in vacations:
            uid = v.user_id
            if uid not in user_dates:
                user_dates[uid] = []
            user_dates[uid].append(v.date)

        def dates_to_periods(dates):
            """Превращает список дат в периоды (start, end, days)."""
            if not dates:
                return []
            sorted_dates = sorted(dates)
            periods = []
            start = sorted_dates[0]
            prev  = sorted_dates[0]
            for d in sorted_dates[1:]:
                if (d - prev).days <= 1:
                    prev = d
                else:
                    periods.append({'start': start.isoformat(), 'end': prev.isoformat(),
                                    'days': (prev - start).days + 1})
                    start = d
                    prev  = d
            periods.append({'start': start.isoformat(), 'end': prev.isoformat(),
                            'days': (prev - start).days + 1})
            return periods

        # Считаем пересечения: для каждой пары дат проверяем overlap
        all_user_dates = {uid: set(user_dates.get(uid, [])) for u in users for uid in [u.id]}
        overlaps = {}  # uid → set of dates where overlap with someone else
        user_ids = [u.id for u in users]
        for i, uid_a in enumerate(user_ids):
            for uid_b in user_ids[i+1:]:
                common = all_user_dates.get(uid_a, set()) & all_user_dates.get(uid_b, set())
                if common:
                    overlaps.setdefault(uid_a, set()).update(common)
                    overlaps.setdefault(uid_b, set()).update(common)

        # Строим ответ
        MONTH_NAMES_RU = ['Январь','Февраль','Март','Апрель','Май','Июнь',
                          'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
        result = []
        for u in users:
            periods = dates_to_periods(user_dates.get(u.id, []))
            total_days = sum(p['days'] for p in periods)

            # Разбиваем периоды по месяцам
            by_month = {m: [] for m in range(1, 13)}
            for p in periods:
                s = dt.date.fromisoformat(p['start'])
                e = dt.date.fromisoformat(p['end'])
                # Период может перекрывать несколько месяцев
                cur = s
                while cur <= e:
                    month_end = dt.date(year, cur.month,
                        [31,28+int(year%4==0),31,30,31,30,31,31,30,31,30,31][cur.month-1])
                    seg_end = min(e, month_end)
                    by_month[cur.month].append({
                        'start': cur.isoformat(),
                        'end': seg_end.isoformat(),
                        'days': (seg_end - cur).days + 1,
                        'has_overlap': bool(overlaps.get(u.id, set()) &
                            {cur + dt.timedelta(days=x) for x in range((seg_end-cur).days+1)}),
                    })
                    cur = seg_end + dt.timedelta(days=1)

            result.append({
                'user_id': u.id,
                'user_name': u.full_name or u.email,
                'total_days': total_days,
                'by_month': by_month,
                'has_overlap': bool(overlaps.get(u.id)),
            })

        return Response({
            'year': year,
            'month_names': MONTH_NAMES_RU,
            'users': result,
        })

    @action(detail=False, methods=['post'], url_path='vacation_export')
    def vacation_export(self, request):
        """Экспорт графика отпусков в Excel — файл или на почту."""
        import datetime as dt, io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from ..models import SystemSettings

        year      = int(request.data.get('year', dt.date.today().year))
        send_via  = request.data.get('send_via', 'file')
        to_email  = request.data.get('to_email', '').strip()

        # Переиспользуем логику vacation_report
        from apps.accounts.models import User
        users = User.objects.filter(is_active=True).exclude(
            role__name='communications'
        ).order_by('last_name', 'first_name')

        vacations = DutySchedule.objects.filter(
            duty_type='vacation', date__year=year
        ).select_related('user').order_by('date')

        user_dates = {}
        for v in vacations:
            user_dates.setdefault(v.user_id, []).append(v.date)

        def dates_to_periods(dates):
            if not dates: return []
            s_dates = sorted(dates)
            periods, start, prev = [], s_dates[0], s_dates[0]
            for d in s_dates[1:]:
                if (d - prev).days <= 1: prev = d
                else:
                    periods.append((start, prev)); start = prev = d
            periods.append((start, prev))
            return periods

        MONTHS = ['Январь','Февраль','Март','Апрель','Май','Июнь',
                  'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

        # Пересечения
        all_ud = {u.id: set(user_dates.get(u.id, [])) for u in users}
        uids = [u.id for u in users]
        overlaps = {}
        for i, a in enumerate(uids):
            for b in uids[i+1:]:
                common = all_ud.get(a, set()) & all_ud.get(b, set())
                if common:
                    overlaps.setdefault(a, set()).update(common)
                    overlaps.setdefault(b, set()).update(common)

        # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Отпуска {year}'

        thin   = Side(style='thin', color='CCCCCC')
        brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
        hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hfill  = PatternFill('solid', start_color='1677FF')
        vfill  = PatternFill('solid', start_color='FFF9C4')   # жёлтый — отпуск
        ofill  = PatternFill('solid', start_color='FFCCBC')   # оранжевый — пересечение
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        calign = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Заголовок
        ws.cell(1, 1, f'График отпусков {year}').font = Font(name='Arial', bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

        # Шапка
        ws.cell(3, 1, 'Сотрудник').font = hfont
        ws.cell(3, 1).fill = hfill
        ws.cell(3, 1).alignment = halign
        ws.cell(3, 1).border = brd
        ws.column_dimensions['A'].width = 26

        for m_idx, m_name in enumerate(MONTHS, 1):
            col = m_idx + 1
            ws.cell(3, col, m_name).font = hfont
            ws.cell(3, col).fill = hfill
            ws.cell(3, col).alignment = halign
            ws.cell(3, col).border = brd
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        ws.cell(3, 14, 'Итого дней').font = hfont
        ws.cell(3, 14).fill = hfill
        ws.cell(3, 14).alignment = halign
        ws.cell(3, 14).border = brd
        ws.column_dimensions['N'].width = 12
        ws.row_dimensions[3].height = 30

        # Данные
        for row_idx, u in enumerate(users, 4):
            periods = dates_to_periods(user_dates.get(u.id, []))
            total   = sum((e-s).days + 1 for s, e in periods)
            has_ov  = bool(overlaps.get(u.id))

            ws.cell(row_idx, 1, u.full_name or u.email)
            ws.cell(row_idx, 1).alignment = calign
            ws.cell(row_idx, 1).border = brd
            ws.cell(row_idx, 1).font = Font(name='Arial', size=10)

            # По месяцам
            by_month_text = {m: [] for m in range(1, 13)}
            by_month_ov   = {m: False for m in range(1, 13)}
            for s, e in periods:
                cur = s
                while cur <= e:
                    month_last = [31,28+int(year%4==0),31,30,31,30,31,31,30,31,30,31][cur.month-1]
                    seg_end = min(e, dt.date(year, cur.month, month_last))
                    days = (seg_end - cur).days + 1
                    by_month_text[cur.month].append(
                        f'{cur.strftime("%d.%m")}-{seg_end.strftime("%d.%m")} ({days}д)'
                    )
                    seg_dates = {cur + dt.timedelta(x) for x in range(days)}
                    if seg_dates & overlaps.get(u.id, set()):
                        by_month_ov[cur.month] = True
                    cur = seg_end + dt.timedelta(days=1)

            for m in range(1, 13):
                col  = m + 1
                text = '\n'.join(by_month_text[m]) if by_month_text[m] else '—'
                cell = ws.cell(row_idx, col, text)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = brd
                cell.font = Font(name='Arial', size=10)
                if by_month_ov[m]:
                    cell.fill = ofill
                elif by_month_text[m]:
                    cell.fill = vfill

            ws.cell(row_idx, 14, total if total else '—')
            ws.cell(row_idx, 14).alignment = halign
            ws.cell(row_idx, 14).border = brd
            ws.cell(row_idx, 14).font = Font(name='Arial', bold=True, size=10)
            ws.row_dimensions[row_idx].height = max(18, 14 * max(
                1, max((len(by_month_text[m]) for m in range(1,13)), default=1)
            ))

        # Легенда
        leg_row = len(list(users)) + 5
        ws.cell(leg_row, 1, '🟡 — отпуск   🟠 — пересечение с другим сотрудником')
        ws.cell(leg_row, 1).font = Font(name='Arial', size=9, italic=True, color='777777')

        ws.freeze_panes = 'B4'
        filename = f'vacation_{year}.xlsx'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        excel_bytes = buf.read()

        if send_via == 'email':
            if not to_email:
                return Response({'error': 'Не указан email'}, status=400)
            s = SystemSettings.get()
            if not s.smtp_host or not s.smtp_user or not s.smtp_password_encrypted or not s.smtp_from_email:
                return Response({'error': 'SMTP настройки не заполнены'}, status=400)
            try:
                import smtplib, ssl
                from email.mime.multipart import MIMEMultipart
                from email.mime.base import MIMEBase
                from email.mime.text import MIMEText
                from email import encoders
                from_addr = f'{s.smtp_from_name} <{s.smtp_from_email}>' if s.smtp_from_name else s.smtp_from_email
                msg = MIMEMultipart()
                msg['Subject'] = f'График отпусков {year}'
                msg['From'] = from_addr
                msg['To'] = to_email
                msg.attach(MIMEText(
                    f'<html><body style="font-family:Arial,sans-serif;padding:20px;">'
                    f'<h2 style="color:#1677ff;">График отпусков {year}</h2>'
                    f'<p>Во вложении файл Excel с графиком отпусков на {year} год.</p>'
                    f'<p style="color:#999;font-size:12px;">Support Portal</p>'
                    f'</body></html>', 'html', 'utf-8'
                ))
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(excel_bytes)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
                if s.smtp_use_ssl:
                    with smtplib.SMTP_SSL(s.smtp_host, s.smtp_port, context=ssl.create_default_context(), timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                elif s.smtp_use_tls:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.ehlo(); srv.starttls(); srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                else:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                return Response({'message': f'Файл отправлен на {to_email}'})
            except Exception as e:
                return Response({'error': f'Ошибка отправки: {e}'}, status=500)

        response = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response



# ===================== ОФД / ККТ =====================
