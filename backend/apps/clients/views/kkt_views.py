import subprocess
import json
import os
from datetime import datetime
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from ..models import Client, ClientNote, CustomFieldDefinition, ClientActivity, Provider, ClientFile, SystemSettings, DutySchedule, CustomHoliday, KktData, OfdCompany
from ..serializers import (
    ClientListSerializer, ClientDetailSerializer, ClientWriteSerializer,
    ClientNoteSerializer, CustomFieldDefinitionSerializer, ProviderSerializer, ClientFileSerializer,
    DutyScheduleSerializer, CustomHolidaySerializer, OfdCompanySerializer, OfdCompanyWriteSerializer,
)
from apps.accounts.permissions import CanEditClient, CanManageCustomFields, IsAdmin
from .utils import ping_ip, build_change_log, FIELD_LABELS, STATUS_LABELS


def parse_datetime(value):
    """Парсит дату из строки ОФД (ISO 8601 без timezone)"""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def ofd_request(inn, token, search=''):
    """
    Один вызов скрипта — получает все ККТ по ИНН, фильтрует по адресу,
    возвращает детали по каждому РНМ. Если search — 16 цифр, делает прямой запрос по РНМ.
    """
    script = '/usr/local/bin/ofd_fetch.sh'
    try:
        args = [script, inn, token, search]
        result = subprocess.run(args, capture_output=True, text=True, timeout=900)
        if not result.stdout.strip():
            stderr = result.stderr.strip()
            raise Exception(f'Пустой ответ от скрипта. stderr: {stderr}')
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        raise Exception(f'Некорректный JSON: {result.stdout[:300]}')
    except subprocess.TimeoutExpired:
        raise Exception('Превышено время ожидания (300 сек)')
    except FileNotFoundError:
        raise Exception(f'Скрипт {script} не найден в контейнере')


def ofd_get_all_rnm(inn, token, timeout=30):
    """
    Прямой запрос к ОФД — возвращает ТОЛЬКО список РНМ без детализации.
    Один запрос на всю компанию — используется для сверки.
    """
    import urllib.request
    import gzip as _gzip
    url = f"https://lk.ofd.ru/api/integration/v2/inn/{inn}/kkts?AuthToken={token}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, */*",
        "Accept-Encoding": "gzip, deflate",
        "Referer": "https://lk.ofd.ru/",
        "Origin": "https://lk.ofd.ru",
    }
    try:
        url.encode("ascii")
    except UnicodeEncodeError:
        raise Exception("Токен содержит недопустимые символы")
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
        if r.headers.get("Content-Encoding") == "gzip":
            raw = _gzip.decompress(raw)
        data = json.loads(raw.decode("utf-8"))
    if data.get("Status") != "Success":
        raise Exception(f'ОФД вернул ошибку: {data}')
    return data.get("Data", [])


class OfdKktView(APIView):
    """
    GET   /api/clients/{id}/ofd_kkt/  — получить сохранённые ККТ из БД
    POST  /api/clients/{id}/ofd_kkt/  — запросить данные с ОФД (полный поиск по ИНН+адресу)
    PATCH /api/clients/{id}/ofd_kkt/  — быстрое обновление по сохранённым РНМ
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, pk=None):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        kkts = client.kkt_data.all()
        data = []
        for k in kkts:
            data.append({
                'id': k.id,
                'kkt_reg_id': k.kkt_reg_id,
                'serial_number': k.serial_number,
                'fn_number': k.fn_number,
                'kkt_model': k.kkt_model,
                'create_date': k.create_date.isoformat() if k.create_date else None,

                'contract_end_date': k.contract_end_date.isoformat() if k.contract_end_date else None,
                'fn_end_date': k.fn_end_date.isoformat() if k.fn_end_date else None,

                'fiscal_address': k.fiscal_address,
                'fetched_at': k.fetched_at.isoformat() if k.fetched_at else None,
            })
        return Response(data)

    def post(self, request, pk=None):
        """Полный поиск: по ИНН компании + фильтрация по адресу аптеки"""
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        if not client.ofd_company:
            return Response({'error': 'У клиента не выбрана компания. Укажите компанию в карточке клиента.'}, status=400)

        address = (client.address or '').strip()
        if not address:
            return Response({'error': 'У клиента не заполнен адрес'}, status=400)

        inn = client.ofd_company.inn.strip()
        token = client.ofd_company.ofd_token.strip()

        if not inn:
            return Response({'error': f'У компании «{client.ofd_company.name}» не заполнен ИНН'}, status=400)
        if not token:
            return Response({'error': f'У компании «{client.ofd_company.name}» не задан токен ОФД. Откройте раздел «Компании» и добавьте токен.'}, status=400)

        # Один вызов скрипта — он сам фильтрует по адресу и получает детали
        try:
            import time
            t_start = time.time()
            result_data = ofd_request(inn, token, address)
            elapsed = round(time.time() - t_start, 1)
        except Exception as e:
            return Response({'error': f'Ошибка запроса к ОФД: {str(e)}'}, status=502)

        if result_data.get('Status') != 'Success':
            return Response({'error': f'ОФД вернул ошибку: {result_data}'}, status=502)

        all_items = result_data.get('Data', [])
        if not all_items:
            return Response({'error': f'ККТ по адресу «{address}» не найдены в ОФД. Проверьте адрес клиента — он должен совпадать с адресом установки кассы в ОФД.'}, status=404)

        fetched = []
        errors = []
        for item in all_items:
            kkt_reg_id = item.get('KktRegId', '')
            kkt_obj, created = KktData.objects.get_or_create(
                client=client,
                kkt_reg_id=kkt_reg_id,
            )
            old_fn = kkt_obj.fn_number  # запоминаем до изменения
            kkt_obj.serial_number = item.get('SerialNumber', '')
            kkt_obj.fn_number = item.get('FnNumber', '')
            kkt_obj.kkt_model = item.get('KktModel', '')
            kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
            kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
            kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
            kkt_obj.fiscal_address = item.get('FiscalAddress', '')
            kkt_obj.raw_data = item
            kkt_obj.save()
            fetched.append(kkt_reg_id)
            if created:
                ClientActivity.objects.create(
                    client=client, user=request.user,
                    action=f'Добавлена ККТ\nРНМ: {kkt_reg_id}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: {kkt_obj.fn_number or "—"}'
                )
            elif old_fn and old_fn != kkt_obj.fn_number:
                ClientActivity.objects.create(
                    client=client, user=request.user,
                    action=f'Изменён номер ФН\nРНМ: {kkt_reg_id}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: «{old_fn}» → «{kkt_obj.fn_number}»'
                )

        return Response({
            'success': True,
            'fetched': fetched,
            'errors': errors,
            'elapsed': elapsed,
            'message': f'Получено ККТ: {len(fetched)}, время: {elapsed} сек' + (f', ошибок: {len(errors)}' if errors else ''),
        })

    def delete(self, request, pk=None, kkt_id=None):
        """DELETE /api/clients/{id}/ofd_kkt/{kkt_id}/ — удалить одну ККТ"""
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        if not kkt_id:
            return Response({'error': 'Не указан ID ККТ'}, status=400)

        try:
            kkt = client.kkt_data.get(pk=kkt_id)
            rnm = kkt.kkt_reg_id
            serial = kkt.serial_number or '—'
            fn_number = kkt.fn_number or '—'
            kkt.delete()
            ClientActivity.objects.create(
                client=client, user=request.user,
                action=f'Удалена ККТ\nРНМ: {rnm}\nСерийный номер: {serial}\nНомер ФН: {fn_number}'
            )
            return Response({'success': True, 'message': 'ККТ удалена'})
        except KktData.DoesNotExist:
            return Response({'error': 'ККТ не найдена'}, status=404)

    def patch(self, request, pk=None):
        """
        Обновление по РНМ.
        Если в теле запроса передан rnm_override — используем его (режим создания клиента).
        Иначе берём РНМ из уже сохранённых в БД записей KktData.
        """
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({'error': 'Клиент не найден'}, status=404)

        # Режим сохранения РНМ без запроса к ОФД (при создании карточки)
        rnm_only_list = request.data.get('rnm_only_list')
        if rnm_only_list:
            import re
            saved = []
            errors = []
            for rnm in rnm_only_list:
                rnm = str(rnm).strip()
                if not rnm:
                    continue
                if not re.match(r'^\d{16}$', rnm):
                    errors.append(f'Некорректный РНМ: «{rnm}». Должно быть 16 цифр.')
                    continue
                kkt_obj, created = KktData.objects.get_or_create(
                    client=client,
                    kkt_reg_id=rnm,
                )
                if created:
                    ClientActivity.objects.create(
                        client=client, user=request.user,
                        action=f'Добавлена ККТ\nРНМ: {rnm}\nСерийный номер: —\nНомер ФН: —'
                    )
                saved.append(rnm)
            return Response({
                'success': True,
                'saved': saved,
                'errors': errors,
                'message': f'Сохранено РНМ: {len(saved)}',
            })

        if not client.ofd_company:
            return Response({'error': 'У клиента не выбрана компания'}, status=400)

        inn = client.ofd_company.inn.strip()
        token = client.ofd_company.ofd_token.strip()

        if not inn:
            return Response({'error': f'У компании «{client.ofd_company.name}» не заполнен ИНН'}, status=400)
        if not token:
            return Response({'error': f'У компании «{client.ofd_company.name}» не задан токен ОФД'}, status=400)

        # Если передан конкретный РНМ — работаем только с ним (для режима создания)
        rnm_override = (request.data.get('rnm_override') or '').strip()
        if rnm_override:
            import re
            if not re.match(r'^\d{16}$', rnm_override):
                return Response({'error': f'Некорректный РНМ: «{rnm_override}». Должно быть 16 цифр.'}, status=400)
            import time
            t_start = time.time()
            script = '/usr/local/bin/ofd_fetch.sh'
            try:
                result = subprocess.run(
                    [script, inn, token, rnm_override],
                    capture_output=True, text=True, timeout=15
                )
                if not result.stdout.strip():
                    return Response({'error': f'Пустой ответ от скрипта для РНМ {rnm_override}. stderr: {result.stderr[:200]}'}, status=502)
                detail_data = json.loads(result.stdout)
            except Exception as e:
                return Response({'error': f'Ошибка запроса для РНМ {rnm_override}: {str(e)}'}, status=502)

            if detail_data.get('Status') != 'Success':
                return Response({'error': f'ОФД вернул ошибку для РНМ {rnm_override}'}, status=502)

            items = detail_data.get('Data', [])
            if not items:
                return Response({'error': f'ОФД не вернул данных для РНМ {rnm_override}'}, status=404)

            # Проверяем что адрес ККТ совпадает с адресом клиента
            client_address = (client.address or '').strip()
            if client_address:
                import re as _re
                stop_words = {'ул', 'пр', 'пом', 'д', 'кв', 'г', 'р-н', 'ул.', 'пр.', 'д.', 'пом.',
                                'г.', 'пр-кт', 'пр-кт.', 'проспект', 'улица', 'переулок', 'помещ',
                                'помещение', 'область', 'край', 'республика', 'город'}
                words = [w.strip('.,()-').lower() for w in _re.split(r'[\s,]+', client_address)]
                words = [w for w in words if w and len(w) > 1 and w not in stop_words]
                house_numbers = _re.findall(r'\b(\d+)\b', client_address)
                house_num = house_numbers[-1] if house_numbers else None

                fiscal_addr = (items[0].get('FiscalAddress') or '').lower()
                addr_numbers = _re.findall(r'\b(\d+)\b', items[0].get('FiscalAddress', ''))

                word_match = all(w in fiscal_addr for w in words)
                house_match = (not house_num) or (house_num in addr_numbers)

                if not (word_match and house_match):
                    return Response({
                        'success': False,
                        'fetched': [],
                        'errors': [
                            f'РНМ {rnm_override}: адрес ККТ в ОФД («{items[0].get("FiscalAddress", "")}») '
                            f'не совпадает с адресом клиента («{client_address}»)'
                        ],
                        'elapsed': round(time.time() - t_start, 1),
                        'message': f'РНМ {rnm_override} не добавлен: адрес не совпадает',
                    }, status=200)  # 200 чтобы фронт показал предупреждение, не ошибку

            for item in items:
                kkt_obj, created = KktData.objects.get_or_create(
                    client=client,
                    kkt_reg_id=item.get('KktRegId', rnm_override),
                )
                old_fn = kkt_obj.fn_number
                kkt_obj.serial_number = item.get('SerialNumber', '')
                kkt_obj.fn_number = item.get('FnNumber', '')
                kkt_obj.kkt_model = item.get('KktModel', '')
                kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
                kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
                kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
                kkt_obj.fiscal_address = item.get('FiscalAddress', '')
                kkt_obj.raw_data = item
                kkt_obj.save()
                if created:
                    ClientActivity.objects.create(
                        client=client, user=request.user,
                        action=f'Добавлена ККТ\nРНМ: {rnm_override}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: {kkt_obj.fn_number or "—"}'
                    )
                elif old_fn and old_fn != kkt_obj.fn_number:
                    ClientActivity.objects.create(
                        client=client, user=request.user,
                        action=f'Изменён номер ФН\nРНМ: {rnm_override}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: «{old_fn}» → «{kkt_obj.fn_number}»'
                    )

            elapsed = round(time.time() - t_start, 1)
            return Response({
                'success': True,
                'fetched': [rnm_override],
                'errors': [],
                'elapsed': elapsed,
                'message': f'Получены данные по РНМ {rnm_override}, время: {elapsed} сек',
            })

        # Стандартный режим — обновляем по РНМ из БД
        kkts = client.kkt_data.all()
        if not kkts:
            return Response({'error': 'Нет сохранённых ККТ. Сначала нажмите «Получить данные с ОФД»'}, status=404)

        import time
        t_start = time.time()
        script = '/usr/local/bin/ofd_fetch.sh'
        fetched = []
        errors = []

        for kkt_obj in kkts:
            rnm = kkt_obj.kkt_reg_id
            if not rnm:
                continue
            try:
                result = subprocess.run(
                    [script, inn, token, rnm],
                    capture_output=True, text=True, timeout=15
                )
                if not result.stdout.strip():
                    errors.append(f'РНМ {rnm}: пустой ответ')
                    continue
                detail_data = json.loads(result.stdout)
            except Exception as e:
                errors.append(f'РНМ {rnm}: {str(e)}')
                continue

            if detail_data.get('Status') != 'Success':
                errors.append(f'РНМ {rnm}: ОФД вернул ошибку')
                continue

            for item in detail_data.get('Data', []):
                old_fn = kkt_obj.fn_number
                old_serial = kkt_obj.serial_number
                was_empty = not old_serial and not old_fn  # запись была только с РНМ
                kkt_obj.serial_number = item.get('SerialNumber', '')
                kkt_obj.fn_number = item.get('FnNumber', '')
                kkt_obj.kkt_model = item.get('KktModel', '')
                kkt_obj.create_date = parse_datetime(item.get('CreateDate'))
                kkt_obj.contract_end_date = parse_datetime(item.get('ContractEndDate'))
                kkt_obj.fn_end_date = parse_datetime(item.get('FnEndDate'))
                kkt_obj.fiscal_address = item.get('FiscalAddress', '')
                kkt_obj.raw_data = item
                kkt_obj.save()
                fetched.append(rnm)
                if was_empty and (kkt_obj.serial_number or kkt_obj.fn_number):
                    # Запись была пустой (только РНМ), теперь заполнена данными с ОФД
                    ClientActivity.objects.create(
                        client=client, user=request.user,
                        action=f'Обновлена ККТ по РНМ\nРНМ: {rnm}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: {kkt_obj.fn_number or "—"}\nМодель: {kkt_obj.kkt_model or "—"}'
                    )
                elif old_fn and old_fn != kkt_obj.fn_number:
                    ClientActivity.objects.create(
                        client=client, user=request.user,
                        action=f'Изменён номер ФН\nРНМ: {rnm}\nСерийный номер: {kkt_obj.serial_number or "—"}\nНомер ФН: «{old_fn}» → «{kkt_obj.fn_number}»'
                    )

        elapsed = round(time.time() - t_start, 1)
        return Response({
            'success': True,
            'fetched': fetched,
            'errors': errors,
            'elapsed': elapsed,
            'message': f'Обновлено ККТ: {len(fetched)}, время: {elapsed} сек' + (f', ошибок: {len(errors)}' if errors else ''),
        })


def _kkt_queryset(params):
    """Общая фильтрация/сортировка KktData по параметрам запроса."""
    from django.db.models import Q
    from ..models import KktData

    qs = KktData.objects.select_related('client', 'client__ofd_company').all()

    search = params.get('search', '').strip() if hasattr(params, 'get') else ''
    if search:
        qs = qs.filter(
            Q(fiscal_address__icontains=search) |
            Q(kkt_reg_id__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(fn_number__icontains=search) |
            Q(client__ofd_company__name__icontains=search) |
            Q(client__address__icontains=search)
        )

    month = params.get('month')
    year  = params.get('year')
    if month and year:
        qs = qs.filter(fn_end_date__year=int(year), fn_end_date__month=int(month))

    allowed_orderings = {
        'client__address', '-client__address',
        'client__ofd_company__name', '-client__ofd_company__name',
        'fn_end_date', '-fn_end_date',
        'contract_end_date', '-contract_end_date',
        'kkt_reg_id', '-kkt_reg_id',
        'serial_number', '-serial_number',
        'fn_number', '-fn_number',
    }
    ordering = params.get('ordering', 'fn_end_date')
    if ordering not in allowed_orderings:
        ordering = 'fn_end_date'
    qs = qs.order_by(ordering)
    return qs


def _kkt_row(k):
    client = k.client
    return {
        'id': k.id,
        'client_id': client.id,
        'address': client.address or k.fiscal_address or '—',
        'company': client.ofd_company.name if client.ofd_company else '—',
        'rnm': k.kkt_reg_id or '—',
        'serial_number': k.serial_number or '—',
        'fn_number': k.fn_number or '—',
        'fn_end_date': k.fn_end_date.isoformat() if k.fn_end_date else None,
        'contract_end_date': k.contract_end_date.isoformat() if k.contract_end_date else None,
    }


class KktListView(APIView):
    """Глобальный список всех ККТ для страницы «Замена ФН»"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = _kkt_queryset(request.query_params)
        page_size = min(int(request.query_params.get('page_size', 50)), 500)
        page      = max(int(request.query_params.get('page', 1)), 1)
        total     = qs.count()
        start     = (page - 1) * page_size
        results   = [_kkt_row(k) for k in qs[start:start + page_size]]
        return Response({'count': total, 'results': results})


class KktExportView(APIView):
    """Экспорт ККТ в Excel — скачать или отправить на почту."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from ..models import SystemSettings
        import datetime

        params    = request.data
        send_via  = params.get('send_via', 'file')
        to_email  = params.get('to_email', '').strip()
        cols_raw  = params.get('cols', '')
        vis_cols  = [c.strip() for c in cols_raw.split(',') if c.strip()] if cols_raw else None

        qs = _kkt_queryset(params)

        COL_DEFS = [
            ('address',           'Адрес',               35, lambda k: k['address']),
            ('company',           'Компания',             25, lambda k: k['company']),
            ('rnm',               'РНМ',                  20, lambda k: k['rnm']),
            ('serial_number',     'Серийный номер',       20, lambda k: k['serial_number']),
            ('fn_number',         'Номер ФН',             20, lambda k: k['fn_number']),
            ('fn_end_date',       'Конец срока ФН',       18, lambda k: k['fn_end_date'][:10] if k['fn_end_date'] else ''),
            ('contract_end_date', 'Конец договора ОФД',  20, lambda k: k['contract_end_date'][:10] if k['contract_end_date'] else ''),
        ]
        headers = [(title, width, getter) for key, title, width, getter in COL_DEFS
                   if vis_cols is None or key in vis_cols]

        rows = [_kkt_row(k) for k in qs]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Замена ФН'

        hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hfill  = PatternFill('solid', start_color='1677FF')
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        calign = Alignment(vertical='center', wrap_text=True)
        thin   = Side(style='thin', color='CCCCCC')
        brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, (title, width, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = brd
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 35

        for ri, row in enumerate(rows, 2):
            fill = PatternFill('solid', start_color='F8FBFF' if ri % 2 == 0 else 'FFFFFF')
            for ci, (_, _, getter) in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=getter(row))
                cell.font = Font(name='Arial', size=10)
                cell.alignment = calign; cell.border = brd; cell.fill = fill
            ws.row_dimensions[ri].height = 18

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(headers))}1'

        filename = f'fn_replacement_{datetime.date.today()}.xlsx'
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        excel_bytes = buf.read()

        if send_via == 'email':
            if not to_email:
                return Response({'error': 'Не указан email'}, status=400)
            s = SystemSettings.get()
            if not s.smtp_host or not s.smtp_user or not s.smtp_password_encrypted or not s.smtp_from_email:
                return Response({'error': 'SMTP настройки не заполнены'}, status=400)
            try:
                import smtplib, ssl
                from email.mime.multipart import MIMEMultipart
                from email.mime.base import MIMEBase
                from email.mime.text import MIMEText
                from email import encoders
                from_addr = f'{s.smtp_from_name} <{s.smtp_from_email}>' if s.smtp_from_name else s.smtp_from_email
                msg = MIMEMultipart()
                msg['Subject'] = f'Замена ФН — {datetime.date.today()}'
                msg['From']    = from_addr
                msg['To']      = to_email
                msg.attach(MIMEText(
                    f'<html><body style="font-family:Arial,sans-serif;padding:20px;">'
                    f'<h2 style="color:#1677ff;">Замена ФН</h2>'
                    f'<p>Во вложении файл Excel с данными по замене ФН на {datetime.date.today()}.</p>'
                    f'<p style="color:#999;font-size:12px;">Support Portal</p>'
                    f'</body></html>', 'html', 'utf-8'
                ))
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(excel_bytes)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
                if s.smtp_use_ssl:
                    ctx = ssl.create_default_context()
                    with smtplib.SMTP_SSL(s.smtp_host, s.smtp_port, context=ctx, timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                elif s.smtp_use_tls:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.ehlo(); srv.starttls(); srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                else:
                    with smtplib.SMTP(s.smtp_host, s.smtp_port, timeout=10) as srv:
                        srv.login(s.smtp_user, s.smtp_password); srv.sendmail(s.smtp_from_email, to_email, msg.as_string())
                return Response({'message': f'Файл отправлен на {to_email}'})
            except Exception as e:
                return Response({'error': f'Ошибка отправки: {e}'}, status=500)

        response = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────
#  Массовый импорт клиентов из JSON
# ─────────────────────────────────────────────